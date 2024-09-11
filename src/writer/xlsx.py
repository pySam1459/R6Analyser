from pathlib import Path
from openpyxl import load_workbook, Workbook
from typing import Optional, TypeAlias

from config import Config
from ignmatrix import IGNMatrix
from history import History, HistoryRound
from utils import transpose, ndefault
from utils.enums import IGNMatrixMode, SaveFileType

from .base import Writer
from .utils import make_copyfile


class XlsxWriter(Writer):
    def __init__(self, save_path: Path, config: Config, append_mode = False) -> None:
        super(XlsxWriter, self).__init__(SaveFileType.XLSX, save_path, config, append_mode)

    def write(self, history: History, ignmat: IGNMatrix) -> None:
        self._pre_write(history, ignmat)

        workbook, save_path, append = self.__load_workbook(self._save_path, self._append_mode)

        ## remove default worksheet
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        ## get data
        rounds = [history.roundn] if append else history.get_round_nums()
        xslx_match = self.__get_xlsx_match(history, ignmat, workbook) if append else self.__get_xlsx_match(history, ignmat)
        data = xslx_match | self.__get_xlsx_rounds(history, rounds)

        ## Add new sheets and fill them with data
        for sheet_name, sheet_data in data.items():
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]

            sheet = workbook.create_sheet(title=sheet_name)
            for row in sheet_data:
                sheet.append(row)
        
        try:
            ## Save the workbook
            workbook.save(save_path)
        except PermissionError:
            new_savepath = make_copyfile()
            workbook.save(new_savepath)
            print(f"SAVE FILE ERROR: Permission Denied! Cannot save to {save_path}, file may already be open. Saving to {new_savepath}")
    
    def __load_workbook(self, save_path: Path, append: bool) -> tuple[Workbook, Path, bool]:
        """Loads a workbook, if `append=True`, it attempts to load the existing workbook from `savepath`"""
        if not append: ## if append=False, return new Workbook object
            return Workbook(), save_path, False

        if not save_path.exists(): ## if save_path does not exist, return new Workbook object
            print(f"SAVE FILE ERROR: {save_path} does not exist! Defaulting to append-save=False.")
            return Workbook(), save_path, False
        
        temp_file = make_copyfile()
        try: ## attempts to load existing workbook
            return load_workbook(save_path), save_path, True
            
        except PermissionError as e:
            print(f"SAVE FILE ERROR: Permission Denied! Cannot open save file {save_path}, file may already be open.\nDefaulting to append-save=False and saving to {temp_file}.\n{str(e)}")
        except Exception as e:
            print(f"SAVE FILE ERROR: An error occurred when trying to load the existing xlsx file {save_path}.\nDefaulting to append-save=False and saving to {temp_file}.\n{str(e)}")

        ## existing workbook failed to load, return a new Workbook object
        return Workbook(), temp_file, False
    

    def __get_xlsx_match(self, history: History, ignmat: IGNMatrix, workbook: Optional[Workbook] = None) -> dict:
        """Returns the data present on the Match sheet, statistics across all rounds"""
        existing_kd = None
        if workbook is not None and "Match" in workbook.sheetnames:
            existing_kd = self.__get_existing_kd(ignmat, workbook)

        headers = [
            ["Statistics"],
            ["Player", "Team Index", "Rounds", "Kills", "Deaths"]]
        
        kills, deaths = self.__get_xlsx_kd(history, existing_kd)
        data = transpose([
            [pl.ign for pl in self._players],
            [ndefault(pl.team, "") for pl in self._players],
            [history.roundn]*len(self._players),
            kills,
            deaths
        ])

        ## if player[0] is not attacker of first recorded game
        assert self._config is not None
        min_rn = min(history.get_round_nums())
        rps = self._config.rounds_per_side
        team = self._players[0].team
        if team is not None:
            if rps < min_rn <= rps*2:
                team = 1-team

            hr = history.get_round(min_rn)
            if hr is not None and hr.atk_side != team:
                data = self.__table_flip(data)

        return { "Match": headers + data }

    KD_TABLE: TypeAlias = tuple[list[int], list[int]]
    def __get_existing_kd(self, ignmat: IGNMatrix, workbook: Workbook) -> KD_TABLE:
        """Returns the kills/deaths from an existing Match sheet"""
        ## TODO: if you are trying to re-append a Round, you will count that round's stat's twice
        existing_kd = [list(row) for row in workbook["Match"].iter_rows(min_row=3, max_row=12, min_col=4, max_col=5, values_only=True)]
        existing_names = next(workbook["Match"].iter_cols(min_col=1, max_col=1, min_row=3, max_row=12, values_only=True))

        if ignmat.mode == IGNMatrixMode.FIXED:
            kills, deaths = transpose([existing_kd[existing_names.index(pl.ign)] for pl in self._players])
            return (kills, deaths)

        ## Existing names must be checked against currently-known names
        pl_map = {} ## maps player.idx to existing_name index
        for i, name in enumerate(existing_names):
            scores = [IGNMatrix.compare_names(pl.ign, str(name)) for pl in self._players] 
            idx = scores.index(max(scores))
            pl_map[self._players[idx].idx] = i

        n_players = len(self._players) ## select existing_kd for new __player list
        kills, deaths = [0]*n_players, [0]*n_players
        for i, pl in enumerate(self._players):
            if pl.idx in pl_map:
                kills[i] = existing_kd[pl_map[pl.idx]][0]
                deaths[i] = existing_kd[pl_map[pl.idx]][1]

        return (kills, deaths)

    def __get_xlsx_kd(self, history: History, existing_kd: Optional[KD_TABLE] = None) -> KD_TABLE:
        """Returns the kills/deaths for all players from the history across all rounds"""
        n_players = len(self._players)
        kills, deaths = ([0]*n_players, [0]*n_players) if existing_kd is None else existing_kd

        idx_map = {pl.idx: i for i, pl in enumerate(self._players)} ## in case player.idx >= 10
        for round in history.get_rounds():
            for record in round.clean_killfeed:
                if record.player.team != record.target.team: ## do not count tk's as a kill
                    kills[idx_map[record.player.idx]] += 1
            for idx in round.clean_deaths:
                deaths[idx_map[idx]] += 1

        return kills, deaths

    def __get_xlsx_rounds(self, history: History, round_nums: list[int]) -> dict[str, list[list]]:
        """Creates a dictionary containing the sheet data for each round"""
        # return {f"Round {rn}": self.__get_xlsx_rdata(round) for rn in round_nums if (round := self.__history.get_round(rn)) is not None}
        data: dict[str, list[list]] = {}
        for rn in round_nums:
            if (round := history.get_round(rn)) is not None:
                data[f"Round {rn}"] = self.__get_xlsx_rdata(round)

        return data

    def __get_xlsx_rdata(self, round: HistoryRound) -> list[list]:
        """This method creates the xlsx data appended to each round sheet, from the data gathered in 1 round"""
        rdata = [
            ["Statistics"],
            ["Player", "Team Index", "Kills", "Deaths", "Assissts", "Hs%", "Headshots", "1vX", "Operator"]
        ]
        ## Statistics Section
        idx_map = {pl.idx: i for i, pl in enumerate(self._players)}
        teams_known = all([pl.team is not None for pl in self._players]) and len(self._players) == 10

        ## calculate the kills/deaths for each player
        n_players = len(self._players)
        kills, deaths = [0]*n_players, [False]*n_players
        round_kf = round.clean_killfeed or round.killfeed
        for record in round_kf:
            if record.player.team != record.target.team:
                kills[idx_map[record.player.idx]] += 1

        for idx in round.clean_deaths:
            deaths[idx_map[idx]] = True

        onevx = None
        onevx_count = 0
        ## TODO: this oneVx calculator won't work for when players and teams are not known
        # your team won, you were the last alive (not necessarily alive at the end), X = number of opponents alive when your last alive teammate died
        if (w := round.winner) is not None and teams_known:
            if deaths[w*5:(w+1)*5].count(False) == 1: ## possible 1vX
                onevx = deaths.index(False)
                for record in reversed(round_kf):
                    if record.player.idx == onevx:
                        onevx_count += 1
                    else:
                        break
                onevx_count += deaths[(1-w)*5:(2-w)*5].count(False)

        ## compile the statistics table for 1 round
        stats_table = []
        for i, pl in enumerate(self._players):
            onevx_pl = 0 if pl.idx != onevx else onevx_count
            stats_table.append([pl.ign, ndefault(pl.team, ""), kills[i], deaths[i], "", "", "", onevx_pl, ""])
        
        ## make sure defence team is at the top of the stats table, same as dissect
        if round.atk_side == self._players[0].team:
            stats_table = self.__table_flip(stats_table)

        rdata += stats_table

        ## Winning team
        prefix = "YOUR TEAM" if round.winner == 0 else "OPPONENTS"
        winning_team_str = f"{prefix} [{round.winner}]"

        ## Round Info
        bpat = round.bomb_planted_at
        ddat = round.disabled_defuser_at
        rdata.extend([
            [],
            ["Round Info"],
            ["Name", "Value", "Time"],
            ["Site"],
            ["Winning team",  winning_team_str],
            ["Win condition", round.win_condition.value]
        ])

        if len(round_kf) == 0:
            rdata.extend([["Opening kill"], ["Opening death"]])
        else:
            opening_kd = round_kf[0]
            rdata.extend([
                ["Opening kill",  opening_kd.player.ign, str(opening_kd.time)],
                ["Opening death", opening_kd.target.ign, str(opening_kd.time)]
            ])
        
        rdata.extend([
            ["Planted at",    str(ndefault(bpat, ""))],
            ["Defused at",    str(ndefault(ddat, ""))],
            [],
            ["Kill/death feed"],
            ["Player", "Target", "Time", "Traded", "Refragged Death", "Refragged Kill"],
        ])

        ## Kill/death feed
        refragged_kills = []
        n_players = len(round_kf)
        for i, record in enumerate(round_kf):
            if i+1 == n_players:
                rdata.append([record.player.ign, record.target.ign, str(record.time), "FALSE", "FALSE", str(i in refragged_kills)])
                break
                
            ## traded = time between this and last kill is <= 6s, the second kill is a player from the opposition
            traded = i+1 < n_players \
                and (record.time - round_kf[i+1].time) <= 6 \
                and record.target.team != round_kf[i+1].target.team
            
            ## refragged = A kills B, A dies in the next 6 seconds
            refragged_death = False
            for j, r2 in enumerate(round_kf[i+1:], start=i+1):
                if record.time - r2.time > 6: break
                if record.player.idx == r2.target.idx:
                    refragged_death = True
                    refragged_kills.append(j)

            rdata.append([record.player.ign, record.target.ign, str(record.time), str(traded), str(refragged_death), str(i in refragged_kills)])
        
        return rdata

    def __table_flip(self, data: list[list]) -> list[list]:
        """Moves the rows 0-4-5-9 to 5-9-0-4"""
        return data[5:] + data[:5]