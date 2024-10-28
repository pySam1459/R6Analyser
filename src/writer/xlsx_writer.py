from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any

from config import Config
from history import History, HistoryRound
from ignmatrix import IGNMatrix, Player
from utils import ndefault, fmt_s
from utils.enums import SaveFileType, Team
from stats import *

from .base import Writer
from .utils import get_players, handle_write_errors


__all__ = ["XlsxWriter"]


MATCH_SHEET_HEADERS = [
    ["Statistics"],
    ["Player", "Team Index", "Rounds", "Kills", "Deaths", "KST", "KPR", "SRV", "Hs%", "1vX"]
]

ROUND_SHEET_STATS_HEADERS = [
    ["Statistics"],
    ["Player", "Team Index", "Kills", "Death", "Assissts", "Hs%", "Headshots", "1vX", "Operator", "Traded Kills", "Refragged Kills", "Traded Deaths"]
]

ROUND_SHEET_INFO_HEADERS = [
    ["Round Info"],
    ["Name", "Value", "Time"]
]

ROUND_SHEET_KILLFEED_HEADERS = [
    ["Kill/death feed"],
    ["Player", "Target", "Time", "Trade", "Refrag"]
]

EMPTY_ROW = [[]]

WINNING_TEAM_TEXT = {
    Team.TEAM0: "YOUR TEAM [0]",
    Team.TEAM1: "OPPONENTS [1]",
    Team.UNKNOWN: ""
}

BOOL_MAP = {
    True: "TRUE",
    False: "FALSE"
}


class XlsxWriter(Writer):
    def __init__(self, save_path: Path, config: Config) -> None:
        super(XlsxWriter, self).__init__(SaveFileType.XLSX, save_path, config)
    
    def write(self, history: History, ignmat: IGNMatrix) -> None:
        if not history.is_ready:
            return

        players = get_players(ignmat, history.get_first_round().atk_side)
        match_stats = compile_match_stats(history, players)

        sheets_data = (self.get_match(match_stats, players) |
                       self.get_rounds(match_stats, history, ignmat, players))

        wb = self.create_workbook()
        self.save_workbook(wb, sheets_data)

    def get_match(self, match_stats: MatchStats_t, players: list[Player]) -> dict[str, Any]:
        match_data = compile_match_summary(match_stats, players)
        xlsx_data = [self.__pms2row(match_data[pl.uid], pl) for pl in players]
        return {"Match": MATCH_SHEET_HEADERS + xlsx_data}
    
    def __pms2row(self, pms: PlayerMatchStats, pl: Player) -> list[Any]:
        ## "Player", "Team Index", "Rounds", "Kills", "Deaths", "KST", "KPR", "SRV", "Hs%", "1vX"
        rnds = pms.rnds
        return [
            pl.ign,
            pl.team.value,
            rnds,
            pms.kills,
            pms.deaths,
            fmt_s(pms.kost,rnds),
            fmt_s(pms.kills,rnds),
            fmt_s(rnds-pms.deaths,rnds),
            fmt_s(pms.hs,pms.kills),
            pms.onevx
        ]
    

    def get_rounds(self, match_stats: MatchStats_t,
                   history: History,
                   ignmat: IGNMatrix,
                   players: list[Player]) -> dict:
        """Creates a dictionary containing the sheet data for each round"""
        return {f"Round {self.__get_roundn(round_stats)}": self.__aggregate_round_data(round_stats, hround, ignmat, players)
                for round_stats, hround in zip(match_stats, history.get_rounds())
                if len(round_stats) > 0}

    def __get_roundn(self, round_stats: RoundStats_t) -> int:
        k0 = next(iter(round_stats))
        return round_stats[k0].rnd

    def __aggregate_round_data(self,
                               round_stats: RoundStats_t,
                               hround: HistoryRound,
                               ignmat: IGNMatrix,
                               players) -> list[list[Any]]:
        return (
            ROUND_SHEET_STATS_HEADERS +
            self.compile_round_statistics(round_stats, hround, ignmat) +
                EMPTY_ROW +
            ROUND_SHEET_INFO_HEADERS +
            self.compile_round_info(hround) + 
                EMPTY_ROW +
            ROUND_SHEET_KILLFEED_HEADERS +
            self.compile_round_killfeed(hround, players)
        )

    def compile_round_statistics(self,
                                 round_stats: RoundStats_t,
                                 hround: HistoryRound,
                                 ignmat: IGNMatrix) -> list[list[Any]]:
        players = get_players(ignmat, hround.atk_side)
        return [self.__compile_player_stats_row(round_stats[pl.uid], pl)
                for pl in players
                if pl.uid in round_stats]

    def __compile_player_stats_row(self, prs: PlayerRoundStats, pl: Player) -> list[Any]:
        ## "Player", "Team Index", "Kills", "Death", "Assissts", "Hs%", "Headshots", "1vX", "Operator", "Traded Kills", "Refragged Kills", "Traded Deaths"
        return [
            pl.ign,
            pl.team,
            prs.kills,
            BOOL_MAP[prs.death == 1],
            "",                             ## cannot track assists
            fmt_s(prs.headshots,prs.kills),
            prs.headshots,
            prs.onevx,
            "",                             ## TODO: track operators
            prs.traded_kills,
            prs.refrag_kills,
            prs.traded_death
        ]

    
    def compile_round_info(self, hround: HistoryRound) -> list[list[Any]]:
        if len(hround.killfeed) == 0:
            okd_pign = okd_tign = okd_time = ""
        else:
            okd = hround.killfeed[0]
            okd_pign, okd_tign, okd_time = okd.player.ign, okd.target.ign, str(okd.time)

        return [
            ["Site"],
            ["Winning team",  WINNING_TEAM_TEXT[hround.winner]], ## TODO: custom team names
            ["Win condition", hround.win_condition.value],
            ["Opening kill",  okd_pign, okd_time],
            ["Opening death", okd_tign, okd_time],
            ["Planted at",    str(ndefault(hround.bomb_planted_at, ""))],
            ["Defused at",    str(ndefault(hround.disabled_defuser_at, ""))],
        ]
    
    def compile_round_killfeed(self, hround: HistoryRound, players: list[Player]) -> list[list[Any]]:
        ## "Player", "Target", "Time", "Trade", "Refrag"
        return [
            [
                record.player.ign,
                record.target.ign,
                str(record.time),
                BOOL_MAP[traded_pl is not None],
                BOOL_MAP[refrag_pl is not None]
            ]
            for record, traded_pl, refrag_pl in iter_killfeed(hround.killfeed, players)]

    ## Save workbook
    def create_workbook(self) -> Workbook:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        return wb
    
    @handle_write_errors
    def save_workbook(self, wb: Workbook, data: dict[str,list]) -> None:
        ## Add new sheets and fill them with data
        for sheet_name, sheet_data in data.items():
            sheet: Worksheet = wb.create_sheet(title=sheet_name)
            for row in sheet_data:
                sheet.append(row)

        wb.save(self._save_path)
