import json
import pyautogui
import cv2
import numpy as np
import easyocr
from string import ascii_letters
from argparse import Namespace
from re import search, fullmatch
from time import time
from os import mkdir
from os.path import join, exists
from sys import exit as sys_exit
from dataclasses import dataclass, field
from abc import ABC, abstractmethod
from enum import Enum
from Levenshtein import ratio as leven_ratio
from openpyxl import load_workbook, Workbook
from typing import TypeAlias, Optional, Any


class StrEnum(Enum):
    """Helper class to implement `from_string` for Enum's with String values"""
    @classmethod
    def from_string(cls, value: str):
        """
        Class method to convert a string to an Enum value, with validity checks.
        """
        for enum_member in cls:
            if enum_member.value == value:
                return enum_member
        raise ValueError(f"'{value}' is not a valid {cls.__name__}")


class IGNMatrixMode(StrEnum):
    """
    Mode for the IGN Matrix
        fixed - all 10 IGNs are present before program starts and remains unchanged throughout
        infer - <10 IGNs are present, the remaining IGNs will be inferred through the killfeed
    """
    FIXED = "fixed"
    INFER = "infer"


@dataclass
class Player:
    """Dataclass containing all necessary player information"""
    idx: int
    ign: str
    team: None | int = None


class IGNMatrix(ABC):
    """
    IGN Matrix infers the true IGN from the EasyOCR reading of the IGN (pseudoIGN/pign) from the killfeed.
    The matrix can be initialised with a prior list of 'fixed' IGNs (IGNs known before starting the game)
    If the matrix is not provided with 10 fixed IGNs, the matrix's output will depend on its initialised mode
    - fixed: will return None for all non-fixed IGNs
    - infer: will infer the non-fixed IGNs from the OCR's output
    The matrix will return an index/ID and the true/most-seen IGN when requested using the `get` method
    The matrix uses Levenshtein distance to determine whether two IGNs are the same (could be improved?)
    """

    @property
    @abstractmethod
    def mode(self) -> IGNMatrixMode:
        ...

    @abstractmethod
    def get(self, pseudoIGN: str, threshold: float = 0.75) -> Optional[Player]:
        ...
    
    @abstractmethod
    def get_from_idx(self, idx: int) -> Optional[Player]:
        ...
    
    @abstractmethod
    def get_team_from_idx(self, idx: int) -> Optional[int]:
        ...
    
    @abstractmethod
    def get_players(self) -> list[Player]:
        """Returns a list of 10 Player objects where each team are grouped in team 0: 0-4 | team 1: 5-9"""
        ...
    
    @abstractmethod
    def evaluate(self, pign: str) -> float:
        ...
    
    def update_team_table(self, pidx: int, tidx: int) -> None:
        ...
    
    @staticmethod
    def new(igns: list[str], mode: IGNMatrixMode) -> 'IGNMatrix':
        """Creates a new IGNMatrix object from a list of fixed IGNs and the IGNMatrix Mode"""
        match mode:
            case IGNMatrixMode.FIXED:
                return IGNMatrixFixed.new(igns)
            case IGNMatrixMode.INFER:
                if len(igns) == 10: return IGNMatrixFixed.new(igns)
                else: return IGNMatrixInfer.new(igns)
            case _:
                raise ValueError(f"Unknown IGNMatrixMode {mode}")

    @staticmethod
    def compare_names(name1: str, name2: str) -> float:
        """Compares two IGN's (pseudo/non-pseudo) using Levenshtein distance, output in the range [0-1]"""
        return leven_ratio(name1.lower(), name2.lower())
    
    @staticmethod
    def _check_fixed(ign_list: list[str], pign: str, threshold: float) -> Optional[Player]:
        """Compares a psuedoIGN to a fixed list of igns and returns a Player object if the pign is in the list"""
        max_score = 0.0
        max_idx   = None
        for i, ign in enumerate(ign_list):
            if ign == pign: return Player(i, ign, int(i >= 5))
            if (score := IGNMatrix.compare_names(pign, ign)) > max_score:
                max_score = score
                max_idx = i

        if max_score > threshold and max_idx is not None:
            return Player(max_idx, ign_list[max_idx], int(max_idx >= 5))
        return None


class IGNMatrixFixed(IGNMatrix):
    """
    This subclass of IGNMatrix only handles the case where all IGNs are known beforehand
    Separating the different IGN modes aims to improve efficiency, readability and modularity
    """
    VALID_THRESHOLD = 0.6 ## threshold for Levenshtein distance to determine equality

    def __init__(self, igns: list[str]) -> None:
        self.__fixmat = igns
    
    @property
    def mode(self) -> IGNMatrixMode:
        return IGNMatrixMode.FIXED
    
    def get(self, pign: str, threshold: float = VALID_THRESHOLD) -> Optional[Player]:
        """Returns the most-likely Player object from a pseudoIGN or None if the pseudoIGN does not meet the threshold"""
        return IGNMatrix._check_fixed(self.__fixmat, pign, threshold)

    def get_from_idx(self, idx: int) -> Optional[Player]:
        """Returns the Player object from their index"""
        if 0 <= idx <= 9:
            return Player(idx, self.__fixmat[idx], int(idx >= 5))

        return None
    
    def get_team_from_idx(self, idx: int) -> Optional[int]:
        """Returns the team index from the player's index"""
        return int(idx >= 5)

    def get_players(self) -> list[Player]:
        """Returns a list of all Players currently playing"""
        return [Player(i, ign, int(i >= 5)) for i, ign in enumerate(self.__fixmat)]

    ## worth adding an cache?
    def evaluate(self, pign: str) -> float:
        """Evaluates a given pseudoIGN and returns as score between 0-1"""
        if pign in self.__fixmat: return 1.0
        return max([IGNMatrix.compare_names(pign, name) for name in self.__fixmat])

    @staticmethod
    def new(igns: list[str]) -> 'IGNMatrixFixed':
        """Type checking of `igns` is done by the `__cparse_IGNS` function, except for length=10 check"""
        if len(igns) != 10:
            raise ValueError(f"Invalid Fixed IGNMatrix argument, must have 10 IGNs, not {len(igns)}")
        
        return IGNMatrixFixed(igns)


class IGNMatrixInfer(IGNMatrix):
    """
    IGN Inference has 3 steps, fixed, semi-fixed, matrix
    1. Fixed      - a pseudoIGN is compared to the list of known IGNs
    2. Semi-Fixed - a pseudoIGN is tested against all of the semi-fixed IGN names-dicts
    3. Matrix     - a pseudoIGN is tested against all other pseudoIGN names-dicts
    When a pseudoIGN is found in either semi-fixed or matrix names-dicts, the occurrence of that pseudoIGN is incremented
    If a matrix names-dict reaches the Assimilation threshold of occurrences, that pseduo IGN names-dict is promoted to semi-fixed
    Once len(Fixed) + len(Semi-Fixed) == 10, no more psuedoIGN names-dict can be promoted and assumed to be invalid IGNs
    A `names-dict: dict[str, int]` is a dictionary containing the occurrency count for each pseudoIGN of an unknown true IGN
    """
    VALID_THRESHOLD = 0.75     ## threshold for Levenshtein distance to determine equality
    ASSIMILATE_THRESHOLD = 7   ## how many times a pseudoIGN has to be seen before adding to 'semi-fixed' list
    KD_DIFF_THRESHOLD = 4      ## how many times player on team 'A' but kill/be killed by a player on team 'B' before setting team
    NO_TEAM_HIST_THRESHOLD = 3 ## how many history records to declare teams, if no team indices are known
    EVAL_DECAY = 0.95

    def __init__(self, igns: list[str]) -> None:
        self.__fixmat = igns
        self.__fixlen = len(igns)

        self.__semi_fixmat: dict[int,dict[str,int]] = {}
        self.__matrix:      dict[int,dict[str,int]] = {}
        self.__semi_fixlen = 0
        self.__idx_counter = self.__fixlen

        self.__team_mat:  dict[int, int]       = {i: int(i>=5) for i in range(self.__fixlen)}
        self.__team_diff: dict[int, list[int]] = {}  ## for each idx, number of kills/deaths to either team
        self.__team_hist: dict[int, list[int]] = {}  ## a list of kill/death relations for each idx
    
    @property
    def mode(self) -> IGNMatrixMode:
        return IGNMatrixMode.INFER
    
    def get(self, pign: str, threshold: float = VALID_THRESHOLD) -> Optional[Player]:
        """
        Returns the most-likely Player from the pseudoIGN
        1. checks the pign against the fixed matrix
        2. checks the pign against the semi-fixed matrix
        3. checks the pign against the matrix & assimmilates player into semi-fixed if assim-threshold is met
        4. otherwise add pign to the matrix
        """
        ## first check the fixed igns
        if self.__fixlen > 0 and (pl := IGNMatrix._check_fixed(self.__fixmat, pign, threshold)) is not None:
            return pl
        
        ## second, check the semi-fixed igns
        for idx, names_dict in self.__semi_fixmat.items():
            if (pl := self.__in_names_dict(pign, idx, names_dict, threshold)):
                return pl
        
        ## if all fixed and semi-fixed igns have been found, assume pseudoIGN is invalid
        if self.__fixlen + self.__semi_fixlen >= 10:
            return None

        ## if not all semi-fixed igns have been found, check matrix, and assimilate if necessary
        for idx, names_dict in self.__matrix.items():
            if (pl := self.__in_names_dict(pign, idx, names_dict, threshold)):
                self.__check_assimilation(idx, names_dict)
                return pl

        ## if ign has not been seen, add to matrix
        idx = self.__idx_counter
        self.__matrix[idx] = { pign: 1 }
        self.__idx_counter += 1
        return Player(idx, pign, None)

    def __in_names_dict(self, pign: str, idx: int, names_dict: dict, threshold: float) -> Optional[Player]:
        """
        This method determines whether a pseudoIGN/pign belongs to a specified `names_dict`, and increases the occurrence counts if it does.
        A `names_dict` is a dictionary with key: value pairs = pign: # occurrences of pign
        """
        if pign in names_dict:
            names_dict[pign] += 1 # increase occurrence count for pseudoIGN
            return Player(idx, IGNMatrixInfer._max_dict(names_dict), self.get_team_from_idx(idx))

        for seen_pign in names_dict.keys():
            if IGNMatrix.compare_names(pign, seen_pign) > threshold:
                names_dict[pign] = 1  # add to names_dict as a possible true IGN
                return Player(idx, IGNMatrixInfer._max_dict(names_dict), self.get_team_from_idx(idx))

        return None
    
    def __check_assimilation(self, idx: int, names_dict: dict[str, int]) -> None:
        """Checks to see if a specified names_dict has passed over the Assimilation threshold, and if so assimilate to semi-fixed matrix"""
        if sum(names_dict.values()) >= IGNMatrixInfer.ASSIMILATE_THRESHOLD:
            self.__semi_fixmat[idx] = self.__matrix.pop(idx)
            self.__semi_fixlen += 1
    
    def get_ign_from_idx(self, idx: int) -> Optional[str]:
        """Returns the player's most-likely ign from their index"""
        if 0 <= idx < self.__fixlen:
            return self.__fixmat[idx]
        elif idx in self.__semi_fixmat:
            return IGNMatrixInfer._max_dict(self.__semi_fixmat[idx])
        elif idx in self.__matrix:
            return IGNMatrixInfer._max_dict(self.__matrix[idx])

        return None

    def get_from_idx(self, idx: int) -> Optional[Player]:
        """Returns the (most likely) Player from a given idx"""
        ign = self.get_ign_from_idx(idx)
        if ign is None:
            return None

        return Player(idx, ign, self.get_team_from_idx(idx))
    
    def get_team_from_idx(self, idx: int) -> Optional[int]:
        """Returns the team index from a player's index"""
        if self.__fixlen >= 5 or idx <= self.__fixlen:
            return int(idx >= 5)

        if idx in self.__team_mat:
            return self.__team_mat[idx]

        return None
    
    def update_team_table(self, pidx: int, tidx: int) -> None:
        """
        To infer a player's team, a record of player kills/deaths are recorded to infer the most likely team they belong to
        The team table is required to handle cases of team-kills, and so the KD_DIFF_THRESHOLD is used to determine 
          when a player has had enough interactions to decide a team
        """
        pteam = self.__team_mat.get(pidx, None)
        tteam = self.__team_mat.get(tidx, None)

        ## parentheses for clarity
        if ((pteam is not None) and (tteam is not None)) \
            or (self.get_team_from_idx(pidx) is not None and self.get_team_from_idx(tidx) is not None): ## both team indices are known
            return

        ## one of the team indices are not known
        elif (pteam is None) and (tteam is not None):
            self.__add_team_diff(pidx, tteam)
        elif pteam is not None and (tteam is None):
            self.__add_team_diff(tidx, pteam)

        ## neither team indices are known
        else:
            self.__add_team_hist(pidx, tidx)
            self.__add_team_hist(tidx, pidx)

    def __add_team_diff(self, idx: int, opp_team: int) -> None:
        """
        This method adds an interaction count with `opp_team` to idx's team_diff
        It then checks to see if `idx`'s team_diff has reached the KD_DIFF_TRESHOLD
          so that idx's team can be set
        """
        if idx not in self.__team_diff:
            self.__team_diff[idx] = [0, 0]
        
        tdiff = self.__team_diff[idx]
        tdiff[opp_team] += 1

        ## check if idx has passed KD_DIFF_THRESHOLD, number of interactions before setting team index
        if abs(tdiff[0] - tdiff[1]) > IGNMatrixInfer.KD_DIFF_THRESHOLD:
            team_idx = int(tdiff[0] > tdiff[1])      ## idx's team = argmin(tdiff)
            self.__set_team_idx(idx, team_idx)

    def __add_team_hist(self, idx: int, value: int) -> None:
        """In the case when neither player's team in an interaction is known, add the interaction to a history"""
        if idx in self.__team_mat: return  ## should only hit if I've made a mistake

        if idx not in self.__team_hist:
            self.__team_hist[idx] = [value]
        else:
            self.__team_hist[idx].append(value)
        
        ## in the case when no player has a set team, declare a team when the # interactions threshold is met
        if len(self.__team_mat) == 0 and len(self.__team_hist[idx]) > IGNMatrixInfer.NO_TEAM_HIST_THRESHOLD:
            self.__set_team_idx(idx, 0)  ## DECLARE TEAMS!

    def __set_team_idx(self, idx: int, team_idx: int) -> None:
        """Sets the team index for a Player with `idx`"""
        self.__team_mat[idx] = team_idx       ## set the team index for idx

        ## use interaction history to infer teams for other players
        if idx in self.__team_hist:
            ## update team_diff for all of idx's previous unknown-team interactions
            for opp_idx in self.__team_hist[idx]:
                self.__add_team_diff(opp_idx, team_idx)
            self.__team_hist.pop(idx)

        if idx in self.__team_diff:
            self.__team_diff.pop(idx)

    def get_players(self) -> list[Player]:
        """Returns a list of Players most likely playing in the game"""
        unsorted_players = self.__get_unsorted_players()
        if self.__fixlen >= 5:
            return unsorted_players

        ## sort the players based on their team, 3rd bucket is for players with an unknown team
        buckets = [[], [], []]
        for pl in unsorted_players:
            bidx = ndefault(pl.team, 2)
            buckets[bidx].append(pl)

        return buckets[0] + buckets[1] + buckets[2]

    def __get_unsorted_players(self) -> list[Player]:
        """
        Creates a list of players with priority starting with fixmat, semi-fixmat and matrix
        The likelihood for players in the matrix is determined by the number of occurrences
        """
        players = [Player(i, ign, int(i >= 5)) for i, ign in enumerate(self.__fixmat)]
        if self.__fixlen == 10:
            return players

        players += [Player(idx, IGNMatrixInfer._max_dict(names_dict), self.get_team_from_idx(idx)) for idx, names_dict in self.__semi_fixmat.items()]
        if len(players) == 10:
            return players
        
        if len(players) + len(self.__matrix) <= 10:
            ## Note: may return a list < 10 length
            return players + [Player(idx, IGNMatrixInfer._max_dict(names_dict), self.get_team_from_idx(idx)) for idx, names_dict in self.__matrix.items()]
        else:
            ## Only add the most likely (most seen) players from __matrix to the players list
            ##   sorted the __matrix by # times seen and only add the top 10-len(players)
            indices = sorted(self.__matrix, key=lambda k: sum(self.__matrix[k].values()), reverse=True)[:10-len(players)]
            return players + [Player(idx, IGNMatrixInfer._max_dict(self.__matrix[idx]), self.get_team_from_idx(idx)) for idx in indices]

    
    def evaluate(self, pign: str) -> float:
        """
        Evaluates an pseudoIGN, determining how close it is to a real ign
        Eval decay: `(IGNMatrixInfer.EVAL_DECAY ** (max_occr - occr)` factor gives emphasis to pseudoIGNs seen more often
        """
        max_val = 0.0
        if pign in self.__fixmat:
            return 1.0

        for ign in self.__fixmat:
            max_val = max(max_val, IGNMatrix.compare_names(pign, ign))
        
        for matrix in [self.__semi_fixmat, self.__matrix]:
            for names_dict in matrix.values():
                max_occr = max(names_dict.values())
                evals = [IGNMatrix.compare_names(pign, name) * (IGNMatrixInfer.EVAL_DECAY ** (max_occr - occr)) for name, occr in names_dict.items()]
                max_val = max(max_val, *evals)

        return max_val

    @staticmethod
    def new(igns: list[str]) -> 'IGNMatrixInfer':
        """Type checking of `igns` is done by the `__cparse_IGNS` function"""
        return IGNMatrixInfer(igns)

    @staticmethod
    def _max_dict(_dict: dict[str, int]) -> str:
        """Returns the key with the maximum value"""
        return max(_dict, key=_dict.get) # type: ignore the line


class TimerFormat(Enum):
    """Enum used to declare what type of formatting the timer requires"""
    FULL    = 0   ## 2:59 - 0:10
    SECONDS = 1   ## 9:99 - 0:00   -> 0:09 - 0:00


@dataclass
class State:
    """Program state, should only be modified by new_round, end_round, fix_state methods"""
    in_round:     bool
    end_round:    bool ## ready's the program to start a new round
    bomb_planted: bool


class WinCondition(StrEnum):
    KILLED_OPPONENTS = "KilledOpponents"
    TIME             = "Time"
    DEFUSED_BOMB     = "DefusedBomb"
    DISABLED_DEFUSER = "DisabledDefuser"
    UNKNOWN          = "Unknown"


@dataclass
class Timestamp:
    """Helper dataclass to deal with the timer"""
    minutes: int
    seconds: int

    def __sub__(self, other: 'Timestamp') -> int:
        return self.to_int() - other.to_int()

    def to_int(self) -> int:
        return self.minutes * 60 + self.seconds

    @staticmethod
    def from_int(num: int) -> 'Timestamp':
        m = int(num / 60)
        s = num - 60*m
        return Timestamp(m, s)
    
    def __str__(self) -> str:
        if self.seconds < 0: return f"{self.minutes}:{self.seconds:03}"
        else: return f"{self.minutes}:{self.seconds:02}"

    __repr__ = __str__


@dataclass
class KFRecord:
    """
    Dataclass to record an player interaction, who killed who, at what time, and if it was a headshot.
      player: killer, target: dead
    """
    player: Player
    target: Player
    time: Timestamp
    headshot: bool = False
    
    def __eq__(self, other: 'KFRecord') -> bool:
        """Equality check only requires index, not time/headshot (A can only kill B once)"""
        return self.player.idx == other.player.idx and self.target.idx == other.target.idx
    
    def update(self, ign_mat: IGNMatrix) -> 'KFRecord':
        """Updates the details for the player/target with the current IGNMatrix data"""
        if ign_mat.mode == IGNMatrixMode.INFER:
            self.player = ign_mat.get_from_idx(self.player.idx) or self.player
            self.target = ign_mat.get_from_idx(self.target.idx) or self.player
        return self

    def to_json(self) -> dict:
        """Converts KFRecord object to json-handlable dictionary"""
        return {
            "time": str(self.time),
            "player": self.player.ign,
            "target": self.target.ign,
            "headshot": self.headshot
        }

    def __str__(self) -> str:
        headshot_str = "(X) " if self.headshot else ""
        return f"{self.time}| {self.player} -> {headshot_str}{self.target}"

    __repr__ = __str__


@dataclass
class HistoryRound:
    """Dataclass storing all of the data gathered from 1 round"""
    scoreline:           Optional[list[int]] = None
    atk_side:            Optional[int] = None
    bomb_planted_at:     Optional[Timestamp] = None
    disabled_defuser_at: Optional[Timestamp] = None
    round_end_at:        Optional[Timestamp] = None
    win_condition:       WinCondition = WinCondition.UNKNOWN
    winner:              Optional[int] = None
    killfeed:            list[KFRecord] = field(default_factory=list)
    deaths:              list[int] = field(default_factory=list)

    clean_killfeed: Optional[list[KFRecord]] = None
    clean_deaths:   Optional[list[int]] = None

    def to_json(self) -> dict:
        """Converts a HistoryRound object to a json-handlable dictionary"""
        return {
            "scoreline":            self.scoreline,
            "atk_side":             self.atk_side,
            "bomb_planted_at":      str(self.bomb_planted_at),
            "disabled_defuser_at":  str(self.disabled_defuser_at),
            "round_end_at":         str(self.round_end_at),
            "win_condition":        self.win_condition.value,
            "winner":               self.winner,
            "killfeed":             [kfr.to_json() for kfr in (self.clean_killfeed if self.clean_killfeed else self.killfeed)]
        }


class History:
    """
    This History class maintains a record of all game data recorded in a dictionary of game round: HistoryRound
    The program's round number counter is coupled to this class and only modified through the `new_round` method
    This class will only record round data after `new_round` is called for the first time
    """
    def __init__(self) -> None:
        self.__roundn = -1
        self.__round_data: dict[int, HistoryRound] = {}
        self.__phantom_round = HistoryRound()

    @property
    def is_ready(self) -> bool:
        return self.__roundn > 0
    
    @property
    def roundn(self) -> int:
        return self.__roundn
    
    def get_round(self, roundn: int) -> Optional[HistoryRound]:
        return self.__round_data.get(roundn, None)
    
    def get_rounds(self) -> list[HistoryRound]:
        return list(self.__round_data.values())
    def get_round_nums(self) -> list[int]:
        return list(self.__round_data.keys())
    def __contains__(self, key: int) -> bool:
        return key in self.__round_data

    @property
    def cround(self) -> HistoryRound:
        """
        Property to access the current HistoryRound
        Returns a Phantom round in-case history is not ready (__roundn <= 0)
        """
        return self.__round_data.get(self.__roundn, self.__phantom_round)

    def new_round(self, round_number: int) -> None:
        """This method should be called at the start of a new round"""
        self.__roundn = round_number
        self.__round_data[round_number] = HistoryRound()
    
    def fix_round(self) -> None:
        """Should be called by _fix_state, in-case program incorrectly thinks round ended"""
        self.cround.round_end_at = None
        
    def to_json(self) -> dict:
        """Converts all game data recorded to json-handlable"""
        return {ridx: round.to_json() for ridx, round in self.__round_data.items()}
    
    def update(self, ignmat: IGNMatrix) -> None:
        """Updates all of the KFRecords with the most up-to-date player information from the IGNMatrix"""
        for round in self.__round_data.values():
            for record in round.killfeed:
                record.update(ignmat)


@dataclass
class SaveFile:
    """Helper dataclass to handle SaveFile paths"""
    filename: str
    ext: str

    def __str__(self) -> str:
        return f"{self.filename}.{self.ext}"
    __repr__ = __str__

    def copy(self) -> 'SaveFile':
        return SaveFile(self.filename, self.ext)


class SaveManager:
    """
    This class manages everything to do with saving round data to json/xlsx
    """
    def __init__(self, savefile: SaveFile, config: dict, append_mode: bool = False) -> None:
        self.__savefile = savefile
        self.__config   = config
        self.__append_mode = append_mode

        self.__players: list[Player]

    def save(self, history: History, ignmat: IGNMatrix) -> None:
        """Main class method to save the history data gathered so-far to json/xlsx"""
        if not exists("saves"):
            mkdir("saves")

        self.__players = ignmat.get_players()
        history.update(ignmat)
        self.__clean_history(history)

        match self.__savefile.ext:
            case "json":
                self.__save_json(history)
            case "xlsx":
                self.__save_xlsx(history, ignmat)
            case _:
                print(f"Unknown save file extension {self.__savefile.ext}")
    
    def __clean_history(self, history: History) -> None:
        """Creates separate clean_killfeed and clean_death attributes to each HistoryRound with all players who aren't in `__players` removed"""
        indices = [pl.idx for pl in self.__players]
        for round in history.get_rounds():
            round.clean_killfeed = [record for record in round.killfeed if record.player.idx in indices and record.target.idx in indices]
            round.clean_deaths   = [didx for didx in round.deaths if didx in indices]

    def __save_json(self, history: History) -> None:
        with open(join("saves", str(self.__savefile)), "w") as f_out:
            json.dump(history.to_json(), f_out, indent=4)


    def __save_xlsx(self, history: History, ignmat: IGNMatrix) -> None:
        savepath = join("saves", str(self.__savefile))
        workbook, savepath, append = self.__load_workbook(savepath, self.__append_mode)

        ## remove default worksheet
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        ## get data
        rounds = [history.roundn] if append else history.get_round_nums()
        xslx_match = self.__get_xlsx_match(history, ignmat, workbook) if append else self.__get_xlsx_match(history, ignmat)
        data = xslx_match | self.__get_xlsx_rounds(history, rounds)

        ## Add new sheets and fill them with data
        for sheet_name, sheet_data in data.items():
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]

            sheet = workbook.create_sheet(title=sheet_name)
            for row in sheet_data:
                sheet.append(row)
        
        try:
            ## Save the workbook
            workbook.save(savepath)
        except PermissionError:
            new_savepath = self.__make_copyfile()
            workbook.save(new_savepath)
            print(f"SAVE FILE ERROR: Permission Denied! Cannot save to {savepath}, file may already be open. Saving to {new_savepath}")
    
    def __load_workbook(self, savepath: str, append: bool) -> tuple[Workbook, str, bool]:
        """Loads a workbook, if `append=True`, it attempts to load the existing workbook from `savepath`"""
        if not append: ## if append=False, return new Workbook object
            return Workbook(), savepath, False

        if not exists(savepath): ## if savepath does not exist, return new Workbook object
            print(f"SAVE FILE ERROR: {savepath} does not exist! Defaulting to append-save=False.")
            return Workbook(), savepath, False
        
        temp_file = self.__make_copyfile()
        try: ## attempts to load existing workbook
            return load_workbook(savepath), savepath, True
            
        except PermissionError as e:
            print(f"SAVE FILE ERROR: Permission Denied! Cannot open save file {savepath}, file may already be open.\nDefaulting to append-save=False and saving to {temp_file}.\n{str(e)}")
        except Exception as e:
            print(f"SAVE FILE ERROR: An error occurred when trying to load the existing xlsx file {savepath}.\nDefaulting to append-save=False and saving to {temp_file}.\n{str(e)}")

        ## existing workbook failed to load, return a new Workbook object
        return Workbook(), temp_file, False
    

    def __get_xlsx_match(self, history: History, ignmat: IGNMatrix, workbook: Optional[Workbook] = None) -> dict:
        """Returns the data present on the Match sheet, statistics across all rounds"""
        existing_kd = None
        if workbook is not None and "Match" in workbook.sheetnames:
            existing_kd = self.__get_existing_kd(ignmat, workbook)

        headers = [
            ["Statistics"],
            ["Player", "Team Index", "Rounds", "Kills", "Deaths"]]
        
        kills, deaths = self.__get_xlsx_kd(history, existing_kd)
        data = transpose([
            [pl.ign for pl in self.__players],
            [ndefault(pl.team, "") for pl in self.__players],
            [history.roundn]*len(self.__players),
            kills,
            deaths
        ])

        ## if player[0] is not attacker of first recorded game
        min_rn = min(history.get_round_nums())
        rps = self.__config["ROUNDS_PER_SIDE"]
        team = self.__players[0].team
        if team is not None:
            if rps < min_rn <= rps*2:
                team = 1-team

            hr = history.get_round(min_rn)
            if hr is not None and hr.atk_side != team:
                data = self.__table_flip(data)

        return { "Match": headers + data }

    def __get_existing_kd(self, ignmat: IGNMatrix, workbook: Workbook) -> tuple[list[int], list[int]]:
        """Returns the kills/deaths from an existing Match sheet"""
        ## TODO: if you are trying to re-append a Round, you will count that round's stat's twice
        existing_kd = [list(row) for row in workbook["Match"].iter_rows(min_row=3, max_row=12, min_col=4, max_col=5, values_only=True)]
        existing_names = next(workbook["Match"].iter_cols(min_col=1, max_col=1, min_row=3, max_row=12, values_only=True))

        if ignmat.mode == IGNMatrixMode.FIXED:
            kills, deaths = transpose([existing_kd[existing_names.index(pl.ign)] for pl in self.__players])
            return (kills, deaths)

        ## Existing names must be checked against currently-known names
        pl_map = {} ## maps player.idx to existing_name index
        for i, name in enumerate(existing_names):
            scores = [IGNMatrix.compare_names(pl.ign, str(name)) for pl in self.__players] 
            idx = scores.index(max(scores))
            pl_map[self.__players[idx].idx] = i

        n_players = len(self.__players) ## select existing_kd for new __player list
        kills, deaths = [0]*n_players, [0]*n_players
        for i, pl in enumerate(self.__players):
            if pl.idx in pl_map:
                kills[i] = existing_kd[pl_map[pl.idx]][0]
                deaths[i] = existing_kd[pl_map[pl.idx]][1]

        return (kills, deaths)

    def __get_xlsx_kd(self, history: History, existing_kd: Optional[tuple[list[int], list[int]]] = None) -> tuple[list[int], list[int]]:
        """Returns the kills/deaths for all players from the history across all rounds"""
        n_players = len(self.__players)
        kills, deaths = ([0]*n_players, [0]*n_players) if existing_kd is None else existing_kd

        idx_map = {pl.idx: i for i, pl in enumerate(self.__players)} ## in case player.idx >= 10
        for round in history.get_rounds():
            assert round.clean_killfeed is not None and round.clean_deaths is not None

            for record in round.clean_killfeed:
                if record.player.team != record.target.team: ## do not count tk's as a kill
                    kills[idx_map[record.player.idx]] += 1
            for idx in round.clean_deaths:
                deaths[idx_map[idx]] += 1

        return kills, deaths

    def __get_xlsx_rounds(self, history: History, round_nums: list[int]) -> dict[str, list[list]]:
        """Creates a dictionary containing the sheet data for each round"""
        # return {f"Round {rn}": self.__get_xlsx_rdata(round) for rn in round_nums if (round := self.__history.get_round(rn)) is not None}
        data: dict[str, list[list]] = {}
        for rn in round_nums:
            if (round := history.get_round(rn)) is not None:
                data[f"Round {rn}"] = self.__get_xlsx_rdata(round)

        return data

    def __get_xlsx_rdata(self, round: HistoryRound) -> list[list]:
        """This method creates the xlsx data appended to each round sheet, from the data gathered in 1 round"""
        rdata = [
            ["Statistics"],
            ["Player", "Team Index", "Kills", "Deaths", "Assissts", "Hs%", "Headshots", "1vX", "Operator"]
        ]
        ## Statistics Section
        idx_map = {pl.idx: i for i, pl in enumerate(self.__players)}
        teams_known = all([pl.team is not None for pl in self.__players]) and len(self.__players) == 10

        ## calculate the kills/deaths for each player
        n_players = len(self.__players)
        kills, deaths = [0]*n_players, [False]*n_players 
        round_kf = ndefault(round.clean_killfeed, round.killfeed)
        for record in round_kf:
            if record.player.team != record.target.team:
                kills[idx_map[record.player.idx]] += 1

        assert round.clean_deaths is not None
        for idx in round.clean_deaths:
            deaths[idx_map[idx]] = True

        onevx = None
        onevx_count = 0
        ## TODO: this oneVx calculator won't work for when players and teams are not known
        # your team won, you were the last alive (not necessarily alive at the end), X = number of opponents alive when your last alive teammate died
        if (w := round.winner) is not None and teams_known:
            if deaths[w*5:(w+1)*5].count(False) == 1: ## possible 1vX
                onevx = deaths.index(False)
                for record in reversed(round_kf):
                    if record.player.idx == onevx:
                        onevx_count += 1
                    else:
                        break
                onevx_count += deaths[(1-w)*5:(2-w)*5].count(False)

        ## compile the statistics table for 1 round
        stats_table = []
        for i, pl in enumerate(self.__players):
            onevx_pl = 0 if pl.idx != onevx else onevx_count
            stats_table.append([pl.ign, ndefault(pl.team, ""), kills[i], deaths[i], "", "", "", onevx_pl, ""])
        
        ## make sure defence team is at the top of the stats table, same as dissect
        if round.atk_side == self.__players[0].team:
            stats_table = self.__table_flip(stats_table)

        rdata += stats_table

        ## Winning team
        prefix = "YOUR TEAM" if round.winner == 0 else "OPPONENTS"
        winning_team_str = f"{prefix} [{round.winner}]"

        ## Round Info
        bpat = round.bomb_planted_at
        ddat = round.disabled_defuser_at
        rdata.extend([
            [],
            ["Round Info"],
            ["Name", "Value", "Time"],
            ["Site"],
            ["Winning team",  winning_team_str],
            ["Win condition", round.win_condition.value]
        ])

        if len(round_kf) == 0:
            rdata.extend([["Opening kill"], ["Opening death"]])
        else:
            opening_kd = round_kf[0]
            rdata.extend([
                ["Opening kill",  opening_kd.player.ign, str(opening_kd.time)],
                ["Opening death", opening_kd.target.ign, str(opening_kd.time)]
            ])
        
        rdata.extend([
            ["Planted at",    str(ndefault(bpat, ""))],
            ["Defused at",    str(ndefault(ddat, ""))],
            [],
            ["Kill/death feed"],
            ["Player", "Target", "Time", "Traded", "Refragged Death", "Refragged Kill"],
        ])

        ## Kill/death feed
        refragged_kills = []
        n_players = len(round_kf)
        for i, record in enumerate(round_kf):
            if i+1 == n_players:
                rdata.append([record.player.ign, record.target.ign, str(record.time), "FALSE", "FALSE", str(i in refragged_kills)])
                break
                
            ## traded = time between this and last kill is <= 6s, the second kill is a player from the opposition
            traded = i+1 < n_players \
                and (record.time - round_kf[i+1].time) <= 6 \
                and record.target.team != round_kf[i+1].target.team
            
            ## refragged = A kills B, A dies in the next 6 seconds
            refragged_death = False
            for j, r2 in enumerate(round_kf[i+1:], start=i+1):
                if record.time - r2.time > 6: break
                if record.player.idx == r2.target.idx:
                    refragged_death = True
                    refragged_kills.append(j)

            rdata.append([record.player.ign, record.target.ign, str(record.time), str(traded), str(refragged_death), str(i in refragged_kills)])
        
        return rdata

    def __make_copyfile(self) -> str:
        """Creates a new savepath based on savefile"""
        temp_savefile = self.__savefile.copy()
        temp_savefile.filename += " - Copy"
        return join("saves", str(temp_savefile))

    def __table_flip(self, data: list[list]) -> list[list]:
        """Moves the rows 0-4-5-9 to 5-9-0-4"""
        return data[5:] + data[:5]


@dataclass
class OCResult:
    """
    A dataclass containing the data of a single easyOCR reading, the data stored includes
        rect - rectangle bounding the text, text, prob - probability assigned by the easyOCR engine, eval_score - IGNMatrix.evaluate score
    """
    rect: list[int]
    text: str
    prob: float
    eval_score: Optional[float] = None

    def __init__(self, ocr_result: tuple[list[list[int]], str, float]):
        self.rect = bbox_to_rect(ocr_result[0])
        self.text = ocr_result[1]
        self.prob = ocr_result[2]

    def eval(self, ign_matrix: IGNMatrix) -> float:
        """Sets the eval_score of self.text given an IGNMatrix"""
        self.eval_score = ign_matrix.evaluate(self.text)
        return self.eval_score
    
    def join(self, other: 'OCResult') -> 'OCResult':
        """
        Combines 2 OCResult objects into 1,
          a solution to a issue with the EasyOCR's detection engine where it wouldn't properly detect a name with an underscore in it
        """
        x, y = min(self.rect[0], other.rect[0]), min(self.rect[1], other.rect[1])
        x2, y2 = max(self.rect[0]+self.rect[2], other.rect[0]+other.rect[2]), max(self.rect[1]+self.rect[3], other.rect[1]+other.rect[3])
        join_box = [[x, y], [x2, y], [x2, y2], [x, y2]]

        join_str  = self.text + "_" + other.text
        join_prob = min(self.prob, other.prob)
        return OCResult((join_box, join_str, join_prob))
    
    def __str__(self) -> str:
        if self.eval_score is not None: return f"{self.text}|eval={self.eval_score}"
        else: return f"{self.text}|prob={self.prob}"

    __repr__ = __str__


@dataclass
class OCRLine:
    """A helper dataclass containing the OCResults and other info from a single killfeed line"""
    ocr_results: list[OCResult]
    headshot: bool = False


class Analyser(ABC):
    """
    Main class `Analyser`
    Operates the main inference loop `run` and records match/round information
    """
    PROB_THRESHOLD = 0.5
    KF_ALLOWLIST = ascii_letters + "0123456789.-_"
    
    def __init__(self, args: Namespace):
        self.config: dict = args.config
        self.verbose: int = args.verbose
        self.prog_args = args
        self._debug_print(f"Config Keys -", list(self.config.keys()))

        if args.check:
            self.check()
            if not args.test: sys_exit()

        self.running = False
        self.tdelta: float = self.config["SCREENSHOT_PERIOD"]

        self.ign_matrix = IGNMatrix.new(self.config["IGNS"], self.config["IGN_MODE"])
        self.reader = easyocr.Reader(['en'], gpu=not args.cpu)
        self._verbose_print(0, "EasyOCR Reader model loaded")

        self.state = State(False, True, False)
        self.history = History()
        self.save_manager = SaveManager(self.prog_args.save, self.config, append_mode=args.append_save)

        self.current_time: Timestamp
        self.defuse_countdown_timer: Optional[float] = None

    ## ----- CHECK -----
    def check(self) -> None:
        """Called when the --check flag is present in the program call, saves the screenshotted regions as jpg images"""
        if not exists("images"):
            mkdir("images")

        self._verbose_print(0, "Saving check images")
        region_keys = self._get_ss_region_keys()
        regions = self._get_ss_regions(region_keys)
        for name, img in regions.items():
            cv2.imwrite(join("images", f"{name}.jpg"), cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
    
    # ----- TEST -----
    @abstractmethod
    def test(self) -> None:
        ...

    # ----- SCREENSHOTS -----
    @abstractmethod
    def _get_ss_region_keys(self) -> list[str]:
        ...

    Regions_t: TypeAlias = dict[str, np.ndarray]
    def _get_ss_regions(self, region_keys: Optional[list[str]] = None) -> Regions_t:
        """Takes a screenshot of the screen, selects regions, and returns them as numpy.ndarray"""
        if region_keys is None:
            region_keys = self._get_ss_region_keys()

        screenshot = pyautogui.screenshot(allScreens=True)
        return {
            region: np.array(screenshot.crop(Analyser.convert_region(self.config[region])), copy=False) # type: ignore the line
            for region in region_keys
        }

    @staticmethod
    def convert_region(region: list[int]) -> tuple[int,int,int,int]:
        """Converts (X,Y,W,H) -> (Left,Top,Right,Bottom)"""
        left, top, width, height = region
        return (left, top, left + width, top + height)

    # ----- MAIN LOOP -----
    def run(self):
        """Main program loop, calls each _handle method every `tdelta` seconds"""
        self.running = True
        self._verbose_print(0, "Running...")

        self.timer = time()
        while self.running:
            if self.timer + self.tdelta > time():
                continue

            __inference_start = time()

            regions = self._get_ss_regions()
            self._handle_scoreline(regions)
            self._handle_timer(regions)
            self._handle_feed(regions)

            self._debug_print(f"Inference time {time()-__inference_start:.2f}s")
            self.timer = time()
    
    ## ----- OCR -----
    def _screenshot_preprocess(self,
                               image: np.ndarray,
                               to_gray: bool = True,
                               denoise: bool = False,
                               squeeze_width: float = -1) -> np.ndarray:
        """
        To increase the accuracy of the EasyOCR readtext function, a few preprocessing techniques are used
          - RGB to Grayscale conversion
          - Denoise the image using fastNlMeansDenoising
          - Resize by factor `Config.SCREENSHOT_RESIZE` (normally 2-4)
          - Squeeze the width of the image, useful for scoreline OCR
        """
        if to_gray:
            image = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)
        if denoise:
            image = cv2.fastNlMeansDenoising(image, None, 5, 7, 21)

        scale_factor = self.config["SCREENSHOT_RESIZE"]
        if squeeze_width != -1:
            sf_w, sf_h = scale_factor * squeeze_width, scale_factor
        else:
            sf_w = sf_h = scale_factor

        new_width = int(image.shape[1] * sf_w)
        new_height = int(image.shape[0] * sf_h)
        return cv2.resize(image, (new_width, new_height), interpolation=cv2.INTER_LINEAR)

    def _readtext(self, image: np.ndarray, prob: float = 0.0, allowlist: Optional[str] = None) -> list[str]:
        """Performs the EasyOCR inference and cleans the output based on the model's assigned probabilities and a threshold"""
        results = self.reader.readtext(image, allowlist=allowlist)
        return [out[1] for out in results if out[2] > prob]


    ## ----- IN ROUND OCR FUNCTIONS -----
    @abstractmethod
    def _handle_scoreline(self, regions: Regions_t) -> None:
        ...
    
    @abstractmethod
    def _handle_timer(self, regions: Regions_t) -> None:
        ...
    
    @abstractmethod
    def _handle_feed(self, regions: Regions_t) -> None:
        ...

    ## ----- GAME STATE FUNCTIONS -----
    @abstractmethod
    def _new_round(self, score1: int, score2: int) -> None:
        ...

    @abstractmethod
    def _end_round(self) -> None:
        ...

    @abstractmethod
    def _fix_state(self) -> None:
        ...    

    # ----- ENG OF PROGRAM / SAVING -----
    @abstractmethod
    def _end_game(self) -> None:
        ...
    
    def _ask_winner(self) -> int:
        """The program cannot currently detect who wins the final round, so get the user to input that info"""
        while (winner := input("Who won the last round? (0/1) >> ")) not in "01":
            print("Invalid option, pick either 0 or 1")
            continue

        return int(winner)

    ## ----- PRINT FUNCTION -----
    def _verbose_print(self, verbose_value: int, *prompt) -> None:
        if self.verbose > verbose_value:
            print("Info:", *prompt)

    def _debug_print(self, *prompt) -> None:
        if self.verbose == 3:
            print("Debug:", *prompt)


class InPersonAnalyser(Analyser):
    NUM_LAST_SECONDS = 4   ## number of seconds to continue reading killfeed after round end (reliability reasons)
    END_ROUND_SECONDS = 12 ## number of seconds to check no timer to determine round end
    
    NUM_KF_LINES = 3
    SCREENSHOT_REGIONS = ["TEAM1_SCORE_REGION", "TEAM2_SCORE_REGION",
                          "TEAM1_SIDE_REGION", "TEAM2_SIDE_REGION",
                          "TIMER_REGION",
                          "KILLFEED_REGION"]
    NON_NAMES = ["has found the bomb", "Friendly Fire has been activated for"]

    SCORELINE_PROB = 0.25
    TIMER_PROB = 0.35
    KF_PROB = 0.15

    PROXIMITY_DIST = 35

    RED_THRESHOLD = 0.73
    RED_RGB_SPACE = np.array([ ## Defines the range for red color in HSV space
        [240, 10, 10],
        [255, 35, 35]])

    def __init__(self, args) -> None:
        super(InPersonAnalyser, self).__init__(args)

        self.last_kf_seconds = None
        self.end_round_seconds = None
        self.last_seconds_count = 0
        self.timer_format = TimerFormat.FULL

        self.atkside_icon = self.__load_side_icon()
        self.kf_lines = self.__get_kf_lines()
    
    def __load_side_icon(self) -> np.ndarray:
        assert exists("res/swords.jpg"), "res/swords.jpg does not exist!"
        img = cv2.imread("res/swords.jpg", cv2.IMREAD_GRAYSCALE)
        side_region_size = self.config["TEAM1_SIDE_REGION"][2:]
        return cv2.resize(img, side_region_size, interpolation=cv2.INTER_LINEAR)

    def _get_ss_region_keys(self) -> list[str]:
        return InPersonAnalyser.SCREENSHOT_REGIONS
    
    def test(self) -> None:
        """
        This method is called when the `--test` flag is added to the program call,
        runs inference on a single screenshot.
        """
        regions   = self._get_ss_regions()
        scoreline = self.__read_scoreline(regions["TEAM1_SCORE_REGION"], regions["TEAM2_SCORE_REGION"])
        atkside   = self.__read_atkside(regions["TEAM1_SIDE_REGION"], regions["TEAM2_SIDE_REGION"])
        time_read = self.__read_timer(regions["TIMER_REGION"])
        feed_read = self.__read_feed(regions["KILLFEED_REGION"])

        print(f"Test: {scoreline=} | {atkside=} | {time_read} | ", end="")
        for line in feed_read:
            if len(line.ocr_results) == 1:
                print(f"{line.ocr_results[0]} -> {line.ocr_results[0]}", end="")
            else:
                headshot = "(X) " if line.headshot else ""
                print(f"{line.ocr_results[0]} -> {headshot}{line.ocr_results[0]}, ", end="")
        print()

    ## ----- SCORELINE -----
    def _handle_scoreline(self, regions: Analyser.Regions_t) -> None:
        """Extracts the current scoreline visible and determines when a new rounds starts"""
        if not self.state.end_round: return

        scores = self.__read_scoreline(regions["TEAM1_SCORE_REGION"], regions["TEAM2_SCORE_REGION"])
        if not scores: return

        new_roundn = sum(scores)+1
        if new_roundn in self.history:
            self._fix_state()
            return

        self._new_round(*scores)

        atkside = self.__read_atkside(regions["TEAM1_SIDE_REGION"], regions["TEAM2_SIDE_REGION"])
        self.history.cround.atk_side = atkside
        self._verbose_print(1, f"Atk Side: {atkside}")
    
    def __read_scoreline(self, scoreline1: np.ndarray, scoreline2: np.ndarray) -> Optional[tuple[int, int]]:
        """Reads the scoreline from the 2 TEAM1/2_SCORE_REGION screenshot"""
        img1 = self._screenshot_preprocess(scoreline1, to_gray=True, squeeze_width=0.65)
        img2 = self._screenshot_preprocess(scoreline2, to_gray=True, squeeze_width=0.65)

        allowlist = "0123456789"
        results = self._readtext(img1, InPersonAnalyser.SCORELINE_PROB, allowlist=allowlist) + self._readtext(img2, InPersonAnalyser.SCORELINE_PROB, allowlist=allowlist)
        if len(results) != 2:
            return None
        if not fullmatch(r"^\d+$", results[0]) or not fullmatch(r"^\d+$", results[1]):
            return None
        
        return (int(results[0]), int(results[1]))
    
    def __read_atkside(self, side1: np.ndarray, side2: np.ndarray) -> Optional[int]:
        """Matches the side icon next to the scoreline with `res/swords.jpg` to determine which side is attack."""
        icon1 = self._screenshot_preprocess(side1, to_gray=True, denoise=True)
        icon2 = self._screenshot_preprocess(side2, to_gray=True, denoise=True)

        res_icon1 = cv2.matchTemplate(self.atkside_icon, icon1, cv2.TM_CCOEFF_NORMED)
        res_icon2 = cv2.matchTemplate(self.atkside_icon, icon2, cv2.TM_CCOEFF_NORMED)

        # Get the maximum match value for each icon
        _, max_val_icon1, _, _ = cv2.minMaxLoc(res_icon1)
        _, max_val_icon2, _, _ = cv2.minMaxLoc(res_icon2)

        # Decide which icon matches best
        if max_val_icon1 >= max_val_icon2:
            return 0
        elif max_val_icon1 < max_val_icon2:
            return 1


    ## ----- TIMER FUNCTION -----
    def _handle_timer(self, regions: Analyser.Regions_t) -> None:
        """Reads and handles the timer information, used to determine when the bomb is planted and when the round ends"""
        if not self.history.is_ready: return

        timer_image = regions["TIMER_REGION"]
        new_time = self.__read_timer(timer_image)

        if new_time is not None: ## timer is showing
            self.current_time = new_time
            self.defuse_countdown_timer = None

            self.last_kf_seconds = None
            self.end_round_seconds = None
        
        elif self.__is_bomb_countdown(timer_image):
            ## The bomb uses a separate countdown timer as the visual timer cannot be accurately tracked
            if self.defuse_countdown_timer is None: ## bomb planted
                self.defuse_countdown_timer = time()
                self.state.bomb_planted = True

                self.history.cround.bomb_planted_at = self.current_time
                self._verbose_print(1, f"Bomb planted at: {self.current_time}")
            elif self.history.cround.bomb_planted_at is not None:
                bpat_int = self.history.cround.bomb_planted_at.to_int()
                self.current_time = Timestamp.from_int(bpat_int - int(time() - self.defuse_countdown_timer))

        elif self.last_kf_seconds is None and self.end_round_seconds is None and self.state.in_round:
            self.last_kf_seconds = time()
            self.end_round_seconds = time()
        
        if self.last_kf_seconds is not None and self.last_kf_seconds + InPersonAnalyser.NUM_LAST_SECONDS < time():
            self.last_kf_seconds = None

        if self.end_round_seconds is not None and self.end_round_seconds + InPersonAnalyser.END_ROUND_SECONDS < time() \
                and self.__read_scoreline(regions["TEAM1_SCORE_REGION"], regions["TEAM2_SCORE_REGION"]) is None:
            self._end_round()
            self.end_round_seconds = None

    
    def __read_timer(self, image: np.ndarray) -> Optional[Timestamp]:
        """
        Reads the current time displayed in the region `TIMER_REGION`
        If the timer is not present, None is returned
        """
        image = self._screenshot_preprocess(image, to_gray=True, denoise=True, squeeze_width=0.75)
        results = self._readtext(image, prob=InPersonAnalyser.TIMER_PROB, allowlist="0123456789:.")
        if len(results) == 0: return None

        result = results[0] if len(results) == 1 else "".join(results)
        time1 = search(r"([0-2]).?([0-5]\d)$", result) ## 2:59 - 0:10
        time2 = search(r"(\d).?\d\d", result)          ## 9:99 - 0:00

        if time1 is None and time2 is not None and self.timer_format != TimerFormat.SECONDS:
            self.last_seconds_count += 1
            if self.last_seconds_count > 4:
                self.timer_format = TimerFormat.SECONDS
                self.last_seconds_count = 0

        if self.timer_format == TimerFormat.FULL:
            if time1 is not None:
                return Timestamp(int(time1.group(1)), int(time1.group(2)))
        elif self.timer_format == TimerFormat.SECONDS or time1 is None:
            if time2 is not None:
                return Timestamp(0, int(time2.group(1)))

        return None

    
    def __is_bomb_countdown(self, image: np.ndarray) -> bool:
        """
        When a bomb is planted, the timer is replaced with a majority red circular countdown
        This method detects when the bomb defuse countdown is shown using a majority red threshold
        """
        mask: np.ndarray = cv2.inRange(image, InPersonAnalyser.RED_RGB_SPACE[0], InPersonAnalyser.RED_RGB_SPACE[1])

        # Calculate the percentage of red in the image
        red_percentage = np.sum(mask > 0) / mask.size
        self._debug_print(f"{red_percentage=}")
        return bool(red_percentage > InPersonAnalyser.RED_THRESHOLD)


    ## ----- KILL FEED -----
    def _handle_feed(self, regions: Analyser.Regions_t) -> None:
        """Handles the killfeed by reading the names, querying the ign matrix and the information to History"""
        if not self.history.is_ready: return
        if not self.state.in_round and self.last_kf_seconds is None: return

        image = regions["KILLFEED_REGION"]
        for line in self.__read_feed(image):
            if len(line.ocr_results) == 1:
                player = self.ign_matrix.get(line.ocr_results[0].text, 0.75)
                target = self.ign_matrix.get(line.ocr_results[0].text, 0.75)

            elif len(line.ocr_results) == 2:
                player = self.ign_matrix.get(line.ocr_results[0].text, 0.75)
                target = self.ign_matrix.get(line.ocr_results[1].text, 0.75)
                
            if player is None or target is None:
                continue ## invalid igns

            record = KFRecord(player, target, self.current_time, line.headshot)
            if record not in self.history.cround.killfeed:
                self.history.cround.killfeed.append(record)
                self.history.cround.deaths.append(target.idx)

                self.ign_matrix.update_team_table(player.idx, target.idx)
                self._verbose_print(1, f"{self.current_time} | {player.ign}\t-> {target.ign}")
    
    def __read_feed(self, image: np.ndarray) -> list[OCRLine]:
        """
        Reading of the killfeed requires a few steps
        1. Preprocessing with grayscale, fastNlMeansDenoising and width squeezing
        2. EasyOCR reading and cleaning
        3. Line Sorting - sorting each OCResult into KF lines
        """
        image = self._screenshot_preprocess(image, to_gray=True, denoise=True, squeeze_width=0.75)

        ocr_results = self.reader.readtext(image, allowlist=Analyser.KF_ALLOWLIST)
        ocr_results = [OCResult(res) for res in ocr_results]
        for res in ocr_results:
            res.eval(self.ign_matrix)

        lines = self.__get_rawlines(ocr_results)
        ocr_lines = self.__get_ocrlines(lines)
        if self.prog_args.test:
            # self.__test_lines(image, lines)
            self.__test_lines(image, ocr_lines)

        return ocr_lines
    
    def __get_rawlines(self, results: list[OCResult]) -> list[list[OCResult]]:
        """This helper method cleans and sorted a raw list of OCResults into a list of NUM_KF_LINES OCRLines"""
        ## clean OCResults by removing prob < KF_PROB, len(text) < 2 and text in NON_NAMES
        temp_results: list[OCResult] = []
        for res in results:
            if res.prob < InPersonAnalyser.KF_PROB or len(res.text) < 2: continue
            if max([IGNMatrix.compare_names(res.text, non_name) for non_name in InPersonAnalyser.NON_NAMES]) > 0.5: continue
            temp_results.append(res)
        
        ## Check which OCResult is in which line using the midline of each OCResult rect
        lines: list[list[OCResult]] = [[] for _ in range(InPersonAnalyser.NUM_KF_LINES)]
        kfr = self.config["KILLFEED_REGION"]
        for res in temp_results:
            midy = (res.rect[1] + res.rect[3]//2) // self.config["SCREENSHOT_RESIZE"]
            for i, lrect in enumerate(self.kf_lines):
                if lrect[1] <= kfr[1]+midy <= lrect[1]+lrect[3]:
                    lines[i].append(res)

        return lines
    
    def __get_ocrlines(self, lines: list[list[OCResult]]) -> list[OCRLine]:
        """
        Converts a list of raw OCR lines into OCRLine objects,
        if more than 2 OCResults are in a line, a join call is required
        """
        ocr_lines = []
        for line in lines:
            if len(line) <= 1:
                ocr_lines.append(OCRLine(line))
            elif len(line) == 2:
                if line[0].rect[0] < line[1].rect[0]:
                    ocr_lines.append(OCRLine(line))
                else:
                    ocr_lines.append(OCRLine([line[1], line[0]]))
            else: ## join call is required
                jline = self.__join_line(line)
                ocr_lines.append(OCRLine(jline))

        return ocr_lines
    
    def __join_line(self, line: list[OCResult]) -> list[OCResult]:
        """Determines which OCResults to join"""
        prox_dist = InPersonAnalyser.PROXIMITY_DIST * (self.config["SCREENSHOT_RESIZE"] / 4)
        line.sort(key=lambda ocr_res: ocr_res.rect[0])
        new_line = []
        used = []
        for i, ocr1 in enumerate(line):
            if i in used: continue
            if i == len(line)-1:
                new_line.append(ocr1)
                break

            for j, ocr2 in enumerate(line[i+1:], start=i+1):
                if j in used: continue
                
                jocr = ocr1.join(ocr2)
                jocr.eval(self.ign_matrix)

                mx_score = max(ocr1.eval_score, ocr2.eval_score) # type: ignore the line
                if jocr.eval_score > mx_score or mx_score == 0.0:
                    new_line.append(jocr)
                elif rect_collision(ocr1.rect, ocr2.rect) or rect_proximity(ocr1.rect, ocr2.rect) <= prox_dist:
                    new_line.append(jocr)
                else:
                    continue

                used += [i, j]
                break
            else:
                new_line.append(ocr1)
                used.append(i)

        if len(new_line) > 2: ## TODO: sort this out
            new_line = sorted(new_line, key=lambda res: res.eval_score, reverse=True)[:2]
            if new_line[0].rect[0] > new_line[1].rect[0]:
                new_line = [new_line[1], new_line[0]]

        return new_line

    def __get_kf_lines(self) -> list[list[int]]:
        kflr = self.config["KF_LINE_REGION"]
        return [[kflr[0], int(kflr[1]-i*1.25*kflr[3]), kflr[2], kflr[3]] for i in range(InPersonAnalyser.NUM_KF_LINES)]


    def __test_lines(self, image: np.ndarray, lines: list[OCRLine]) -> None:
        rect_img = cv2.cvtColor(image.copy(), cv2.COLOR_GRAY2BGR)
        line_cols = [(0, 0, 255), (0, 255, 0), (255, 0, 0)]
        for line, col in zip(lines, line_cols):
            for ores in line.ocr_results:
                cv2.rectangle(rect_img, ores.rect[:2],  [ores.rect[0]+ores.rect[2], ores.rect[1]+ores.rect[3]], col, 3)
        cv2.imwrite("images/TEST_KFLINE_BOXES.jpg", rect_img)

    # ----- GAME STATE -----
    def __pre_new_round(self, score2: int, save: bool = True) -> None:
        """
        To save the correct information for each round, this method must be called just before the start of a new round,
          so that the scoreline and winner attributes can be used
        """
        if not self.history.is_ready: return

        ## infer winner of previous round based on new scoreline
        winner = self.history.cround.winner
        if winner is None and self.history.cround.scoreline is not None:
            _, _score2 = self.history.cround.scoreline
            winner = int(_score2 < score2)
            self.history.cround.winner = winner ## if _score1+1 == score1, return 0

        win_con = self._get_wincon()
        self.history.cround.win_condition = win_con

        reat = self.history.cround.round_end_at
        if win_con == WinCondition.DISABLED_DEFUSER:
            self.history.cround.disabled_defuser_at = reat
            self._verbose_print(1, f"Disabled defsuer at: {reat}")
        
        self._verbose_print(0, f"Team {winner} wins round {self.history.roundn} by {win_con} at {reat}.")
        if save:
            self.save_manager.save(self.history, self.ign_matrix)

    def _new_round(self, score1: int, score2: int) -> None:
        """
        When a new round starts, this method is called, initialising a new round history
        The parameters `score1` and `score2` are the current scores displayed at the start of a new round
        """
        new_round = score1 + score2 + 1
        if new_round in self.history: return

        self.__pre_new_round(score2, save=True)
        self.state = State(True, False, False)
        self.timer_format = TimerFormat.FULL
        self.last_seconds_count = 0

        self.history.new_round(new_round)
        self.history.cround.scoreline = [score1, score2]
        self._verbose_print(1, f"New Round: {new_round} | Scoreline: {score1}-{score2}")
    
    def _end_round(self) -> None:
        """A game state method called when the program determines that the current round has ended"""
        self.history.cround.round_end_at = self.current_time
        self.state = State(False, True, False)
        self.last_seconds_count = 0

        mx_rnd = self.config["MAX_ROUNDS"]
        rps = self.config["ROUNDS_PER_SIDE"]
        scoreline = self.history.cround.scoreline = [0, 0]
        is_ot = not self.config["SCRIM"] and scoreline[0] >= rps and scoreline[1] >= rps
        if self.history.roundn >= mx_rnd \
                or (not is_ot and max(scoreline) == rps+1) \
                or (is_ot and abs(scoreline[0]-scoreline[1]) == 2):
            self._end_game()
    
    def _get_wincon(self) -> WinCondition:
        """Returns the win condition from the round history"""
        bpat    = self.history.cround.bomb_planted_at
        winner  = self.history.cround.winner
        atkside = self.history.cround.atk_side
        if winner is None or atkside is None:
            return WinCondition.UNKNOWN

        defside = 1-atkside
        if bpat is not None and winner == defside:
            return WinCondition.DISABLED_DEFUSER
        elif bpat is not None:
            return WinCondition.DEFUSED_BOMB

        elif 0 <= self.current_time.to_int() <= 1 \
                and winner == defside \
                and self.__wincon_alive_count(atkside) > 0 \
                and self.__wincon_alive_count(defside) > 0:
            return WinCondition.TIME

        elif self.__wincon_alive_count(1-winner) == 0:
            return WinCondition.KILLED_OPPONENTS
        
        return WinCondition.UNKNOWN
    
    def __wincon_alive_count(self, side: int) -> int:
        """Returns the number of alive players on a particular side"""
        alive = 5 ## TODO: history, ign matrix team count
        for d_idx in self.history.cround.deaths:
            if self.ign_matrix.get_team_from_idx(d_idx) == side:
                alive -= 1
        return alive
    
    def _end_game(self) -> None:
        """This method is called when the program determines the game has ended"""
        winner = self._ask_winner()
        self.history.cround.winner = winner
        if self.history.cround.scoreline is not None:
            scoreline = self.history.cround.scoreline[:]
            scoreline[winner] += 1
            self.__pre_new_round(scoreline[1], save=False)

        self.save_manager.save(self.history, self.ign_matrix)
        self._verbose_print(0, f"Data Saved to {self.prog_args.save}, program terminated.")
        sys_exit()
    
    def _fix_state(self) -> None:
        """
        Called when the program incorrectly thinks the round ended, e.g. paused during death animation with now scoreline showing
        """
        self._verbose_print(1, f"Fixing State")
        self.state.in_round = True
        self.state.end_round = False

        timer_region = self._get_ss_regions(["TIMER_REGION"])["TIMER_REGION"]
        self.state.bomb_planted = self.__is_bomb_countdown(timer_region)
        
        self.history.fix_round()


class SpectatorAnalyser(Analyser):
    NUM_LAST_SECONDS = 4
    END_ROUND_SECONDS = 12
    
    SCREENSHOT_REGIONS = ["TEAM1_SCORE_REGION", "TEAM2_SCORE_REGION", "TIMER_REGION", "KILLFEED_REGION"]

    def __init__(self, args) -> None:
        super(SpectatorAnalyser, self).__init__(args)

    def _get_ss_region_keys(self) -> list[str]:
        return SpectatorAnalyser.SCREENSHOT_REGIONS
    
    def test(self) -> None:
        ...

    ## ----- IN ROUND OCR FUNCTIONS -----
    def _handle_scoreline(self, team1_scoreline: np.ndarray, team2_scoreline: np.ndarray) -> None:
        ...
    
    def _handle_timer(self, timer: np.ndarray) -> None:
        ...
    
    def _handle_feed(self, feed: np.ndarray) -> None:
        ...

    ## ----- GAME STATE FUNCTIONS -----
    def _new_round(self, score1: int, score2: int) -> None:
        ...

    def _end_round(self) -> None:
        ...    
    
    def _end_game(self) -> None:
        ...
    
    def _fix_state(self) -> None:
        ...


# ----- HELPER FUNCTIONS -----
def ndefault(value, default):
    if value is None: return default
    return value

def transpose(matrix: list[list[Any]]) -> list[list[Any]]:
    return [list(row) for row in zip(*matrix)]


def rect_collision(r1: list[int], r2: list[int]) -> bool:
    return r1[0] <= r2[0]+r2[2] and r1[1] <= r2[1]+r2[3] \
            and r2[0] <= r1[0]+r1[2] and r2[1] <= r1[1]+r1[3]


def bbox_to_rect(bbox: list[list[int]]) -> list[int]:
    tl, tr, _, bl = bbox
    return [int(tl[0]), int(tl[1]), int(tr[0]-tl[0]), int(bl[1]-tl[1])]


def box_collision(b1: list[list[int]], b2: list[list[int]]) -> bool:
    return rect_collision(bbox_to_rect(b1), bbox_to_rect(b2))


def rect_proximity(r1: list[int], r2: list[int]) -> float:
    if not (r1[1] <= r2[1]+r2[3]/2 <= r1[1]+r1[3]): return float("inf")
    return r2[0] - (r1[0]+r1[2])


def point_in_rect(point: list[int], rect: list[int]) -> bool:
    return rect[0] <= point[0] <= rect[0]+rect[2] and rect[1] <= point[1] <= rect[1]+rect[3]
